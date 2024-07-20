import tkinter as tk
from tkinter import filedialog
from tkcalendar import DateEntry  # Importing DateEntry for date selection
import openpyxl
import reportlab
from PyPDF2 import PdfWriter, PdfReader
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A5

class InputForm(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Excel Data Processor")
        self.geometry("500x250")  # Reduced height to half
        self.configure(bg="white")  # Set background color
        self.resizable(False, False)  # Disable window resizing

        self.date = None
        self.excel_file_path = "C:/Users/Biswarup Roy/Downloads/data.xlsx"  # Default file path
        self.selected_sheet = None
        self.additional_info = None

        self.create_widgets()

    def create_widgets(self):
        # Main Frame to fill entire window
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Excel Frame
        self.frame_excel = tk.Frame(self, bg="white")
        self.frame_excel.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        tk.Label(self.frame_excel, text="Excel File:", bg="white").grid(row=0, column=0, padx=10, pady=10, sticky="ns")
        self.excel_file_entry = tk.Entry(self.frame_excel, width=30)
        self.excel_file_entry.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        self.excel_file_entry.insert(0, self.excel_file_path)  # Set default file path
        browse_button = tk.Button(self.frame_excel, text="Browse", command=self.browse_excel_file)
        browse_button.grid(row=0, column=2, padx=10, pady=10, sticky="ns")
        next_excel_button = tk.Button(self.frame_excel, text="Next", command=self.show_sheet_frame)
        next_excel_button.grid(row=1, column=2, padx=10, pady=10, sticky="se")
        back_excel_button = tk.Button(self.frame_excel, text="Back", command=self.show_date_frame)
        back_excel_button.grid(row=1, column=1, padx=10, pady=10, sticky="sw")

        # Sheet Frame
        self.frame_sheet = tk.Frame(self, bg="white")
        tk.Label(self.frame_sheet, text="Select Sheet:", bg="white").grid(row=0, column=0, padx=10, pady=10, sticky="ns")
        self.sheet_var = tk.StringVar(self.frame_sheet)
        self.sheet_var.set("")  # Initial value
        self.sheet_dropdown = tk.OptionMenu(self.frame_sheet, self.sheet_var, "")
        self.sheet_dropdown.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        back_sheet_button = tk.Button(self.frame_sheet, text="Back", command=self.show_excel_frame)
        back_sheet_button.grid(row=1, column=1, padx=10, pady=10, sticky="sw")
        next_sheet_button = tk.Button(self.frame_sheet, text="Next", command=self.show_date_frame)
        next_sheet_button.grid(row=1, column=2, padx=10, pady=10, sticky="se")

        # Date Frame
        self.frame_date = tk.Frame(self, bg="white")
        tk.Label(self.frame_date, text="Select Bill Date:", bg="white").grid(row=0, column=0, padx=10, pady=10, sticky="ns")
        self.date_entry = DateEntry(self.frame_date, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-mm-yyyy')
        self.date_entry.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        tk.Label(self.frame_date, text="Bill for the Month of:", bg="white").grid(row=1, column=0, padx=10, pady=10, sticky="ns")
        self.additional_info_entry = tk.Entry(self.frame_date, width=30)
        self.additional_info_entry.grid(row=1, column=1, padx=10, pady=10, sticky="ew")
        back_date_button = tk.Button(self.frame_date, text="Back", command=self.show_sheet_frame)
        back_date_button.grid(row=2, column=1, padx=10, pady=10, sticky="sw")
        submit_button = tk.Button(self.frame_date, text="Submit", command=self.submit_form)
        submit_button.grid(row=2, column=2, padx=10, pady=10, sticky="se")

        # Processing Frame
        self.frame_processing = tk.Frame(self, bg="white")
        self.processing_label = tk.Label(self.frame_processing, text="Processing, please wait...", bg="white")
        self.processing_label.pack(padx=20, pady=20)

        # Show initial frame
        self.show_frame(self.frame_excel)

    def show_frame(self, frame):
        self.frame_excel.grid_forget()
        self.frame_sheet.grid_forget()
        self.frame_date.grid_forget()
        self.frame_processing.grid_forget()

        frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        frame.tkraise()

    def show_excel_frame(self):
        self.show_frame(self.frame_excel)

    def show_sheet_frame(self):
        self.show_frame(self.frame_sheet)

        self.clear_sheet_dropdown()

        if self.excel_file_path:
            wb = openpyxl.load_workbook(self.excel_file_path)
            sheet_names = wb.sheetnames
            self.sheet_dropdown = tk.OptionMenu(self.frame_sheet, self.sheet_var, *sheet_names)
            self.sheet_dropdown.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
            wb.close()

            # Set default sheet selection
            if sheet_names:
                # Only set default if it's not already selected
                if not self.sheet_var.get() or self.sheet_var.get() not in sheet_names:
                    self.sheet_var.set(sheet_names[0])

    def show_date_frame(self):
        self.show_frame(self.frame_date)

    def browse_excel_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if file_path:
            self.excel_file_path = file_path
            self.excel_file_entry.delete(0, tk.END)
            self.excel_file_entry.insert(0, self.excel_file_path)

            # Move to sheet selection frame after selecting the excel file
            self.show_sheet_frame()

            # Reset sheet dropdown
            self.set_default_sheet()

    def set_default_sheet(self):
        # Set default sheet selection
        if self.excel_file_path:
            try:
                wb = openpyxl.load_workbook(self.excel_file_path)
                sheet_names = wb.sheetnames
                wb.close()

                if sheet_names:
                    # Only set default if it's not already selected
                    if not self.sheet_var.get() or self.sheet_var.get() not in sheet_names:
                        self.sheet_var.set(sheet_names[0])
            except Exception as e:
                print(f"Error loading workbook: {e}")

    def submit_form(self):
        self.date = self.date_entry.get_date()
        self.selected_sheet = self.sheet_var.get()
        self.additional_info = self.additional_info_entry.get()

        # Show processing frame
        print("processing frame start")
        self.show_frame(self.frame_processing)

        # Process and close after a delay
        self.after(100,self.process_and_close)

    def process_and_close(self):
        print("process function start")
        
        # Example of what you might do with the inputs:
        # formatted_date = self.date.strftime("%d-%m-%Y")  # Format date as date-month-year
        # print("Date:", formatted_date)
        # print("Excel File:", self.excel_file_path)
        # print("Selected Sheet:", self.selected_sheet)
        # print("Additional Info:", self.additional_info)
        
        
        
        packet = io.BytesIO()
        self.create_overlap_all_pages(packet)
        
        
        packet.seek(0)
        new_pdf = PdfReader(packet)
        
        
        
        output = PdfWriter()
        
        for i in range(len(new_pdf.pages)):
          existing_pdf = PdfReader(open("C:\\Users\\Biswarup Roy\\Downloads\\blue-template.pdf", "rb"))
          page = existing_pdf.pages[0]
          page.merge_page(new_pdf.pages[i])
          output.add_page(page)
        
        
        outputStream = open("C:\\Users\\Biswarup Roy\\Downloads\\destination.pdf", "wb")
        output.write(outputStream)
        outputStream.close()
        
        
        # Destroy the processing frame after 3 seconds
        self.after(30, self.destroy_processing_frame)
        

        

    def destroy_processing_frame(self):
        self.frame_processing.grid_forget()

        # Close the application
        self.destroy()

    def clear_sheet_dropdown(self):
        # Clear existing OptionMenu contents
        menu = self.sheet_dropdown['menu']
        menu.delete(0, 'end')
        
    def convert_to_words(self, num):
        if num < 0:
            return "Please Don't Pay This Bill"
    
        ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
        tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
        teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen",
                 "Nineteen"]
    
        words = ""
    
    
        if num >= 1000:
            words += self.convert_to_words(num // 1000) + " Thousand "
        else:
          words += "Rs. "
          
        if num == 0:
          words += "Zero"
          
        num %= 1000
        if num >= 100:
            words += ones[num // 100] + " Hundred "
        num %= 100
        if 10 <= num <= 19:
            words += teens[num - 10] + " "
            num = 0
        elif num >= 20:
            words += tens[num // 10] + " "
        num %= 10
        if 1 <= num <= 9:
            words += ones[num] + " "
        
        return words.strip()

    
    def read_row(self, titles, row):
      entry = {"charges":{}}
    
      for i in range(len(titles)):
        # if row[i].value == None: continue
        if row[i].value == None: row[i].value = 0

        match titles[i].value:
          case "Sl. No.":             key = "id"
          case "Name of Members":     key = "name"
          case "Flat No.":            key = "flat"
          case "MEMBERSHIP NO.":      key = "member"
          case "PREVIOUS DUES":       key = "dues"
          case "ADVANCE RECEIVED":    key = "advance"
          case "SUB TOTAL" | "TOTAL": continue
          case _:
            if row[i].value == 0: continue
            entry["charges"][titles[i].value] = row[i].value
            continue
    
        entry[key] = row[i].value
    
      return entry
      
    
    def read_data(self, file_name, sheet_name):
      wb = openpyxl.load_workbook(file_name)
      sheet = wb[sheet_name]
    
      data = []
      titles = sheet.iter_rows(0,1)
      titles = [i for i in titles]
      titles = titles[0]
    
    
      # for row in range(0,sheet.max_row):
      # for row in sheet.iter_rows(2, 6):
      for row in sheet.iter_rows(2, sheet.max_row):
        # for col in range(0,sheet.max_column):
          data.append(self.read_row(titles, row))
          # print(row[col].value)
      return data
      
    
    def create_overlap_single_page(self, canvas, content):
      canvas.drawString(80, 402, content["name"])
      canvas.drawString(88, 430, content["flat"])
      canvas.drawString(280, 430, str(content["member"]))
      canvas.drawString(84 , 458, self.additional_info[0:3] + "/" + self.additional_info[-4:] + "/" + str(content["id"]))
      canvas.drawString(310, 458, self.date.strftime("%d-%m-%Y"))
      canvas.drawString(225, 486, self.additional_info)  
    
      charges = content["charges"]
      total = 0
      for i,item in enumerate(charges.keys()):
        # print(charges[item])
        total += charges[item]
        canvas.drawString(45,  350 - i * 21, item)
        canvas.drawString(345, 350 - i * 21, (((6 - len(str(charges[item]))) * 2) * " ") + str(charges[item]))
    
      if content.get("dues") != 0:
          total += content["dues"]
          canvas.drawString(45, 140, "Previous Dues")
          canvas.drawString(345, 140, (((6 - len(str(content["dues"]))) *2 ) * " ") + str(content["dues"]))
    
      if content.get("advance") != 0:
           total += content["advance"]
           canvas.drawString(45, 140, "Advance Paid")
           canvas.drawString(345, 140, (((6 - len(str(content["advance"]))) * 2) * " ") + str(content["advance"]))
    
      canvas.drawString(345, 120, (((6 - len(str(total))) * 2) * " ") + str(total))
      canvas.drawString(40, 100, self.convert_to_words(total) + ("" if total < 0 else " Only"))
      
      
    def create_overlap_all_pages(self, overlap_filename):
      # List of content for each page
        data = self.read_data(self.excel_file_path, self.selected_sheet)
    
        # Create a PDF canvas
        c = canvas.Canvas(overlap_filename, pagesize=A5)
    
        # Set font and size
        c.setFont("Helvetica", 12)
    
        # Loop through page contents
        for i, content in enumerate(data):
            # Start a new page (except for the first page)
            if i > 0:
                c.showPage()
    
            # Draw content on the page
            self.create_overlap_single_page(c, content)
    
        # Save the PDF
        c.save()
  
if __name__ == "__main__":
    app = InputForm()
    app.mainloop()
